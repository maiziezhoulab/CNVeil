import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def process_act_excel(input_dir, act_excel, out_dir):
    # Step 1: Load mean ploidy values from TN1–TN8
    mean_ploidy_values = []
    for i in range(1, 9):
        file_path = os.path.join(input_dir, f'TN{i}', 'mean_ploidy')
        with open(file_path, 'r') as f:
            lines = f.readlines()
            last_line = lines[-1]
            mean_ploidy = round(float(last_line.strip().split('Overall ploidy: ')[-1]),2)
            mean_ploidy_values.append(mean_ploidy)

    # Step 2: Load ACT Excel
    df = pd.read_excel(act_excel)

    # Step 3: Add CNVeil column
    df['CNVeil'] = mean_ploidy_values

    # Step 4: Create workbook and sheet
    wb = load_workbook(act_excel)
    ws = wb.active

    # Clear existing rows except header
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Define colors
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Step 5: Write updated DataFrame with conditional formatting
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

        if r_idx == 1:
            continue  # skip header

        # Compute abs diff
        facs_val = df.loc[r_idx - 2, 'FACS']
        tool_vals = df.columns[2:]  # All tools (including CNVeil)
        diffs = {tool: abs(df.loc[r_idx - 2, tool] - facs_val) for tool in tool_vals}
        sorted_tools = sorted(diffs.items(), key=lambda x: x[1])
        
        # Highlight lowest and second lowest diffs
        tool_to_color = {
            sorted_tools[0][0]: green_fill,
            sorted_tools[1][0]: yellow_fill
        }

        for tool, fill in tool_to_color.items():
            col_idx = df.columns.get_loc(tool) + 1
            ws.cell(row=r_idx, column=col_idx).fill = fill

    # Save output
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'ACT.xlsx')
    wb.save(out_path)
    print(f"Saved to {out_path}")

# Example usage:
in_dir = "./"
act_excel="/data/maiziezhou_lab/CanLuo/Single_Cell_Project/DEV/Data/ACT.xlsx"
out_dir = "./evaluation"

process_act_excel(in_dir, act_excel, out_dir)
