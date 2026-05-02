import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def collect_cnveil_values(input_dir):
    sections = ['A', 'B', 'C', 'D','E']
    cnveil_values = []
    for sec in sections:
        file_path = os.path.join(input_dir, f'S0_section_{sec}', f'CNV_S0_section_{sec}.csv.T')
        df = pd.read_csv(file_path, sep=',')
        df_values = df.iloc[:, 1:]  # Exclude first column
        mean_val = round(df_values.values.mean(),2)
        n_cell = df_values.shape[1]
        cnveil_values.extend([mean_val, n_cell])
    return cnveil_values

def update_10x_excel(input_dir, tenx_excel, out_dir):
    cnveil_values = collect_cnveil_values(input_dir)

    # Load 10x Excel file
    df = pd.read_excel(tenx_excel)

    # Add CNVeil column
    df['CNVeil'] = cnveil_values

    # Define fills
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Write to Excel using openpyxl
    out_path = os.path.join(out_dir, '10x.xlsx')
    os.makedirs(out_dir, exist_ok=True)

    # Save with pandas first
    df.to_excel(out_path, index=False)

    # Reopen with openpyxl to apply formatting
    wb = load_workbook(out_path)
    ws = wb.active

    # Identify columns
    header = [cell.value for cell in ws[1]]
    ground_truth_col = header.index("10x summary") + 1

    tool_cols = [i + 1 for i, name in enumerate(header) if i > 2]  # skip Dataset and 10x summary

    # Apply conditional formatting row by row
    for row_idx in range(2, ws.max_row + 1):
        if row_idx%2:
            continue
        gt_val = ws.cell(row=row_idx, column=ground_truth_col).value
        diffs = {}
        for col_idx in tool_cols:
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                diffs[col_idx] = abs(val - gt_val)
        sorted_diffs = sorted(diffs.items(), key=lambda x: x[1])
        if len(sorted_diffs) > 0:
            ws.cell(row=row_idx, column=sorted_diffs[0][0]).fill = green_fill
        if len(sorted_diffs) > 1:
            ws.cell(row=row_idx, column=sorted_diffs[1][0]).fill = yellow_fill

    wb.save(out_path)
    print(f"Saved to: {out_path}")

in_dir = "./"
out_dir = './evaluation'
tenx_excel = "/data/maiziezhou_lab/CanLuo/Single_Cell_Project/DEV/Data/10x.xlsx"
update_10x_excel(in_dir, tenx_excel, out_dir)
